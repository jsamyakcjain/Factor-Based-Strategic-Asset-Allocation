"""
Factor-Based SAA Analytics Report
Run from project root: python generate_analytics.py
Outputs: analytics_report.xlsx
"""
from __future__ import annotations
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.gridspec import GridSpec
import io
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ── Load pipeline ─────────────────────────────────────────────────
print("Loading data and running models...")
from data.data_manager import DataManager
from factors.factor_model import OLSFactorModel, QuantileFactorModel
from models.covariance import POETCovariance
from models.returns import ExpectedReturns
from portfolio.mvo import MVO
from portfolio.risk_parity import RiskParity
from portfolio.hrp import EnhancedHRP
from portfolio.risk_decomp import FactorRiskDecomposition
from config.settings import ASSET_DISPLAY_NAMES

def dn(a): return ASSET_DISPLAY_NAMES.get(a, a.replace("_"," ").title())

dm = DataManager(use_cache=True); dm.load_cached()
ols = OLSFactorModel(dm.factor_returns_t1, dm.asset_returns_t1, credit_liquidity=dm.credit_liquidity)
result = ols.fit()
poet   = POETCovariance(dm.factor_returns_t1, dm.asset_returns_t1_complete, result.betas)
poet.fit()
cov    = poet.as_dataframe()
F      = dm.factor_returns_t1.astype(float)
factor_cov = pd.DataFrame(np.cov(F.values.T), index=F.columns, columns=F.columns)
er     = ExpectedReturns(assets=list(cov.index)); mu = er.quarterly()
mvo_w  = MVO(mu, cov).fit()
rp_w   = RiskParity(cov).fit()
hrp_w  = EnhancedHRP(cov, result.betas).fit()
rd     = FactorRiskDecomposition(result.betas, factor_cov, cov)
assets     = list(cov.index)
eq_assets  = [a for a in assets if a in ["us_large_cap","us_mid_cap","us_small_cap","em_equity","reits"]]
bnd_assets = [a for a in assets if a in ["long_treasury","tips","ig_credit"]]
portfolios = {
    "Equal Weight": rd.equal_weight(assets),
    "60/40":        rd.sixty_forty(eq_assets, bnd_assets),
    "MVO":          mvo_w,
    "Risk Parity":  rp_w,
    "Enhanced HRP": hrp_w,
}
decomp = rd.compare(portfolios)
qm = QuantileFactorModel(dm.factor_returns_t1, dm.asset_returns_t1); qm.fit()
print("Models complete.")

# ── Style helpers ─────────────────────────────────────────────────
NAVY   = "0A1628"; BLUE  = "1A3A5C"; ACCENT = "2E6DA4"
WHITE  = "F0F4F8"; MUTED  = "7A90A8"; PANEL  = "16202E"
RED    = "B03A2E"; AMBER  = "B7770D"; GREEN  = "1E7B45"
LGREY  = "D0D8E4"

PORT_HEX = {
    "Equal Weight": "7F8C8D",
    "60/40":        "B03A2E",
    "MVO":          "2980B9",
    "Risk Parity":  "E67E22",
    "Enhanced HRP": "1E7B45",
}
FACTOR_HEX = {
    "Equity Premium": "C0392B",
    "Term Premium":   "2980B9",
    "Credit Spread":  "E67E22",
    "Inflation":      "27AE60",
    "Liquidity":      "8E44AD",
    "Idiosyncratic":  "566573",
}

def fill(hex_): return PatternFill("solid", start_color=hex_, end_color=hex_)
def font(hex_="F0F4F8", sz=10, bold=False):
    return Font(name="Arial", color=hex_, size=sz, bold=bold)
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border_thin():
    s = Side(style="thin", color="1E2D3D")
    return Border(left=s, right=s, top=s, bottom=s)
def header_row(ws, row, labels, widths=None, bg=NAVY):
    for c, label in enumerate(labels, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.fill = fill(bg); cell.font = font("F0F4F8", 10, True)
        cell.alignment = aln("center"); cell.border = border_thin()
        if widths: ws.column_dimensions[get_column_letter(c)].width = widths[c-1]
def data_row(ws, row, values, bg=PANEL, fmt_map=None):
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill(bg); cell.font = font("D0D8E4", 10)
        cell.alignment = aln("center"); cell.border = border_thin()
        if fmt_map and c in fmt_map: cell.number_format = fmt_map[c]
def section_title(ws, row, col, text, span=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = fill(BLUE); cell.font = font("F0F4F8", 11, True)
    cell.alignment = aln("left"); cell.border = border_thin()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)

def img_to_xl(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight",
                facecolor="#0e1117", edgecolor="none")
    buf.seek(0); plt.close(fig)
    return XLImage(buf)

# Chart style
def fig_style():
    plt.rcParams.update({
        "figure.facecolor":  "#0e1117",
        "axes.facecolor":    "#16202E",
        "axes.edgecolor":    "#1E2D3D",
        "axes.labelcolor":   "#D0D8E4",
        "xtick.color":       "#D0D8E4",
        "ytick.color":       "#D0D8E4",
        "text.color":        "#D0D8E4",
        "grid.color":        "#1E2D3D",
        "grid.linewidth":    0.6,
        "font.family":       "Arial",
        "font.size":         9,
    })

fig_style()

# ── Pre-compute factor column keys (systematic factors only, no alpha) ──
factor_col_keys = list(factor_cov.index)   # e.g. ["equity_premium","term_premium",...]

# 1. Identify the 60/40 Weights
w_6040 = portfolios["60/40"]

# 2. Align weights with asset returns (matching your port_metrics logic)
asset_rets = dm.asset_returns_t1_complete.dropna()
w_al_6040 = w_6040.reindex(asset_rets.columns).fillna(0)
if w_al_6040.sum() != 0:
    w_al_6040 = w_al_6040 / w_al_6040.sum()

# 3. Define the benchmark return series (The 60/40 performance)
benchmark_ret = asset_rets @ w_al_6040

# ── Portfolio metrics helper ──────────────────────────────────────
def port_metrics(w):
    wv = w.reindex(cov.index).fillna(0).values
    if wv.sum() != 0:
        wv = wv / wv.sum()
    mu_v = mu.reindex(cov.index).fillna(0).values

    ann_ret = float(wv @ mu_v) * 4 * 100
    ann_vol = float(np.sqrt(wv @ cov.values @ wv)) * 2 * 100   # *2 == sqrt(4) for quarterly->annual
    sharpe = ann_ret / ann_vol if ann_vol > 0 else 0

    asset_rets_clean = dm.asset_returns_t1_complete.dropna()

    w_al = w.reindex(asset_rets_clean.columns).fillna(0)
    if w_al.sum() != 0:
        w_al = w_al / w_al.sum()

    pr = asset_rets_clean @ w_al

    cum = (1 + pr).cumprod()
    roll_max = cum.cummax()
    dd = (cum - roll_max) / roll_max
    max_dd = float(dd.min()) * 100

    calmar = ann_ret / abs(max_dd) if max_dd < 0 else 0

    # Sortino — use full-sample semideviation (set positives to 0), not subset std
    excess_down = np.minimum(pr.values, 0.0)
    downside_vol = float(np.std(excess_down) * np.sqrt(4) * 100)
    sortino = ann_ret / downside_vol if downside_vol > 0 else 0

    diff = pr - benchmark_ret
    tracking_error = float(np.std(diff) * np.sqrt(4) * 100)

    return ann_ret, ann_vol, sharpe, sortino, max_dd, calmar, tracking_error, pr

# Compute once
metrics = {n: port_metrics(w) for n, w in portfolios.items()}

# ── Mean annualised factor returns (used in alpha decomp) ─────────
# factor_col_keys are the systematic factor columns (no alpha intercept)
mean_quarterly_factor_rets = {
    f: float(dm.factor_returns_t1[f].mean())
    for f in factor_col_keys
    if f in dm.factor_returns_t1.columns
}

# ── Stress periods ────────────────────────────────────────────────
stress_periods = {
    "GFC (2008 Q3 - 2009 Q1)": ("2008-07-01", "2009-03-31"),
    "COVID Crash (2020 Q1)":   ("2020-01-01", "2020-06-30"),
    "Rate Shock (2022)":       ("2022-01-01", "2022-12-31"),
}

def period_return(port_ret, start, end):
    mask = (port_ret.index >= pd.Timestamp(start)) & (port_ret.index <= pd.Timestamp(end))
    sub  = port_ret[mask]
    if len(sub) == 0: return float("nan")
    return float((1 + sub).prod() - 1) * 100

# ═════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ═════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

print("Building Excel workbook...")

# ═══ SHEET 1: EXECUTIVE SUMMARY ═══════════════════════════════════
ws1 = wb.create_sheet("Executive Summary")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 32
for col in ["B","C","D","E","F"]: ws1.column_dimensions[col].width = 18

# Title block
ws1.merge_cells("A1:F1")
c = ws1["A1"]; c.value = "FACTOR-BASED STRATEGIC ASSET ALLOCATION — ANALYTICS REPORT"
c.fill = fill(NAVY); c.font = font("F0F4F8", 14, True)
c.alignment = aln("center")

ws1.merge_cells("A2:F2")
c = ws1["A2"]; c.value = "POET Covariance  |  JPMorgan 2026 LTCMA  |  Enhanced HRP  |  13 Asset Classes  |  2004 Q1 – 2024 Q4"
c.fill = fill(BLUE); c.font = font(MUTED, 10)
c.alignment = aln("center")

ws1.row_dimensions[3].height = 8

# Key finding
ws1.merge_cells("A4:F4")
erp_60 = float(decomp.loc["60/40","Equity Premium"]) if "60/40" in decomp.index else 0
erp_hrp = float(decomp.loc["Enhanced HRP","Equity Premium"]) if "Enhanced HRP" in decomp.index else 0
c = ws1["A4"]
c.value = (f"Key Finding: A conventional 60/40 portfolio allocates {erp_60:.1f}% of total risk to Equity Premium. "
           f"Enhanced HRP reduces this to {erp_hrp:.1f}% — a {erp_60-erp_hrp:.1f} pp reduction — "
           f"while maintaining balanced systematic risk exposure across all five factors.")
c.fill = fill(BLUE); c.font = font("F0F4F8", 10)
c.alignment = aln("left", wrap=True)
ws1.row_dimensions[4].height = 36

ws1.row_dimensions[5].height = 8

# Portfolio metrics table
section_title(ws1, 6, 1, "Portfolio Performance Metrics", 8)
header_row(ws1, 7,
    ["Portfolio","Ann. Return","Ann. Volatility","Sharpe Ratio","Sortino Ratio","Max Drawdown","Calmar Ratio","Tracking Error"],
    [32, 18, 18, 18, 18, 18, 18, 18])
bg_alt = [PANEL, "111A27"]
for i, (name, w) in enumerate(portfolios.items()):
    ann_ret, ann_vol, sharpe, sortino, max_dd, calmar, tracking_error, _ = metrics[name]
    bg = bg_alt[i % 2]
    data_row(ws1, 8+i, [name, ann_ret/100, ann_vol/100, sharpe, sortino, max_dd/100, calmar, tracking_error/100],
             bg=bg, fmt_map={2:"0.0%",3:"0.0%",4:"0.00",5:"0.00",6:"0.0%",7:"0.00",8:"0.0%"})
    ws1.cell(8+i, 1).font = font("F0F4F8", 10, True)
    ws1.cell(8+i, 1).alignment = aln("left")

ws1.row_dimensions[13].height = 8

# Factor attribution summary
section_title(ws1, 14, 1, "Factor Risk Attribution (% of Total Portfolio Risk)", 6)
factor_cols = [c for c in ["Equity Premium","Term Premium","Credit Spread","Inflation","Liquidity","Idiosyncratic"]
               if c in decomp.columns]
header_row(ws1, 15, ["Portfolio"] + factor_cols, [32]+[18]*len(factor_cols))
for i, (name, _) in enumerate(portfolios.items()):
    if name not in decomp.index: continue
    row_data = [name] + [decomp.loc[name, c]/100 for c in factor_cols]
    bg = bg_alt[i % 2]
    fmt = {j+2: "0.0%" for j in range(len(factor_cols))}
    data_row(ws1, 16+i, row_data, bg=bg, fmt_map=fmt)
    ws1.cell(16+i, 1).font = font("F0F4F8", 10, True)
    ws1.cell(16+i, 1).alignment = aln("left")
    # Color ERP cell
    erp_val = decomp.loc[name,"Equity Premium"] if "Equity Premium" in decomp.columns else 0
    erp_cell = ws1.cell(16+i, 2)
    if erp_val > 70: erp_cell.font = Font(name="Arial", color=RED, size=10, bold=True)
    elif erp_val > 55: erp_cell.font = Font(name="Arial", color=AMBER, size=10, bold=True)
    else: erp_cell.font = Font(name="Arial", color=GREEN, size=10, bold=True)

ws1.row_dimensions[21].height = 8

# POET diagnostics
# Filter betas to factor columns only (exclude alpha) for variance decomposition
B_factors = result.betas.reindex(columns=factor_col_keys).values.astype(float)
total_var  = np.trace(cov.values)
factor_var = np.trace(B_factors @ factor_cov.values @ B_factors.T)
sys_share  = factor_var / total_var
idio_share = 1 - sys_share

section_title(ws1, 22, 1, "POET Covariance Diagnostics", 4)
header_row(ws1, 23, ["Parameter","Value","Description"], [32,18,40])
T = len(F)
p = cov.shape[0]
k = F.shape[1]

diag_rows = [
    ("Sample Observations (T)", T, "Complete-case balanced panel"),
    ("Asset Classes (p)", p, "Number of assets"),
    ("Systematic Factors (k)", k, "Factor model dimension"),
    ("Systematic Variance Share", f"{sys_share:.1%}", "Variance explained by factors"),
    ("Idiosyncratic Share", f"{idio_share:.1%}", "Residual variance"),
]
for i, (param, val, desc) in enumerate(diag_rows):
    bg = bg_alt[i % 2]
    ws1.cell(24+i, 1, param).fill = fill(bg)
    ws1.cell(24+i, 1).font = font("F0F4F8", 10)
    ws1.cell(24+i, 1).border = border_thin()
    ws1.cell(24+i, 1).alignment = aln("left")
    ws1.cell(24+i, 2, val).fill = fill(bg)
    ws1.cell(24+i, 2).font = font("D0D8E4", 10)
    ws1.cell(24+i, 2).border = border_thin()
    ws1.cell(24+i, 2).alignment = aln("center")
    ws1.cell(24+i, 3, desc).fill = fill(bg)
    ws1.cell(24+i, 3).font = font(MUTED, 9)
    ws1.cell(24+i, 3).border = border_thin()
    ws1.cell(24+i, 3).alignment = aln("left")


# ═══ SHEET: BACKTEST SUMMARY ══════════════════════════════════════
ws_bt = wb.create_sheet("Backtest Summary", 1)
ws_bt.sheet_view.showGridLines = False
ws_bt.column_dimensions["A"].width = 28
for col in ["B","C","D","E","F","G","H"]:
    ws_bt.column_dimensions[col].width = 18

section_title(ws_bt, 1, 1, "Backtest Summary Statistics — All Portfolios (2004 Q4 – 2024 Q4)", 8)
header_row(ws_bt, 2,
    ["Portfolio","Ann. Return","Ann. Volatility","Sharpe Ratio",
     "Sortino Ratio","Max Drawdown","Calmar Ratio","Tracking Error vs 60/40"],
    [28,16,18,16,16,16,16,26])

for i, (name, _) in enumerate(portfolios.items()):
    ann_ret, ann_vol, sharpe, sortino, max_dd, calmar, te, pr = metrics[name]
    bg = bg_alt[i % 2]
    data_row(ws_bt, 3+i,
             [name, ann_ret/100, ann_vol/100, sharpe, sortino, max_dd/100, calmar, te/100],
             bg=bg, fmt_map={2:"0.0%",3:"0.0%",4:"0.00",5:"0.00",6:"0.0%",7:"0.00",8:"0.0%"})
    ws_bt.cell(3+i, 1).font = font("F0F4F8", 10, True)
    ws_bt.cell(3+i, 1).alignment = aln("left")
    # Color Sharpe
    sharpe_cell = ws_bt.cell(3+i, 4)
    if sharpe >= 0.7: sharpe_cell.font = Font(name="Arial", color=GREEN, size=10, bold=True)
    elif sharpe >= 0.4: sharpe_cell.font = Font(name="Arial", color=AMBER, size=10)
    else: sharpe_cell.font = Font(name="Arial", color=RED, size=10)
    # Color Max DD
    dd_cell = ws_bt.cell(3+i, 6)
    if max_dd < -25: dd_cell.font = Font(name="Arial", color=RED, size=10, bold=True)
    elif max_dd < -15: dd_cell.font = Font(name="Arial", color=AMBER, size=10)
    else: dd_cell.font = Font(name="Arial", color=GREEN, size=10)

# Add rolling Sharpe chart
fig, ax = plt.subplots(figsize=(11, 4))
window = 8  # 8 quarters = 2 years
for name, _ in portfolios.items():
    pr = metrics[name][-1]
    roll_ret = pr.rolling(window).mean() * 4 * 100
    roll_vol = pr.rolling(window).std() * np.sqrt(4) * 100
    roll_sharpe = roll_ret / roll_vol
    ax.plot(roll_sharpe.index, roll_sharpe.values,
            label=name, color="#"+PORT_HEX[name], linewidth=1.5)
ax.axhline(y=0, color="#7A90A8", linewidth=0.8, linestyle="--")
ax.axhline(y=0.5, color="#1E7B45", linewidth=0.8, linestyle=":", alpha=0.6, label="0.5 reference")
ax.set_title("Rolling 2-Year Sharpe Ratio — All Portfolios", fontsize=11,
             fontweight="bold", color="#F0F4F8")
ax.set_ylabel("Sharpe Ratio")
ax.legend(fontsize=9, framealpha=0.3)
ax.grid(alpha=0.4)
img = img_to_xl(fig)
img.width = 770; img.height = 300
ws_bt.add_image(img, f"A{4+len(portfolios)+2}")

# Active Share section
row_as = 4 + len(portfolios) + 20

section_title(ws_bt, row_as, 1, "Active Share vs 60/40 Benchmark", 4)
header_row(ws_bt, row_as+1,
    ["Portfolio","Tracking Error (ann.)","Active Share","Classification"],
    [28, 22, 18, 30])

w_bench = portfolios["60/40"]
for i, (name, w) in enumerate(portfolios.items()):
    _, _, _, _, _, _, te, _ = metrics[name]
    all_assets = list(set(list(w.index) + list(w_bench.index)))
    w1 = w.reindex(all_assets).fillna(0)
    w2 = w_bench.reindex(all_assets).fillna(0)
    active_share = float(0.5 * np.sum(np.abs(w1.values - w2.values)))
    if active_share < 0.2: classification = "Closet Indexer"
    elif active_share < 0.5: classification = "Moderate Active"
    elif active_share < 0.8: classification = "High Conviction Active"
    else: classification = "Very High Active"
    bg = bg_alt[i % 2]
    data_row(ws_bt, row_as+2+i,
             [name, te/100, active_share, classification],
             bg=bg, fmt_map={2:"0.0%", 3:"0.0%"})
    ws_bt.cell(row_as+2+i, 1).font = font("F0F4F8", 10, True)
    ws_bt.cell(row_as+2+i, 1).alignment = aln("left")
    ws_bt.cell(row_as+2+i, 4).alignment = aln("left")
    as_cell = ws_bt.cell(row_as+2+i, 3)
    if active_share > 0.5: as_cell.font = Font(name="Arial", color=GREEN, size=10, bold=True)
    elif active_share > 0.2: as_cell.font = Font(name="Arial", color=AMBER, size=10)
    else: as_cell.font = Font(name="Arial", color=RED, size=10)

# ═══ SHEET 2: PORTFOLIO CHARTS ════════════════════════════════════
ws2 = wb.create_sheet("Portfolio Analytics")
ws2.sheet_view.showGridLines = False
section_title(ws2, 1, 1, "Portfolio Analytics — Charts", 1)
ws2.column_dimensions["A"].width = 2
ws2.column_dimensions["B"].width = 2

port_names = list(portfolios.keys())
port_colors = ["#"+PORT_HEX[n] for n in port_names]
factor_cols_all = [c for c in ["Equity Premium","Term Premium","Credit Spread","Inflation","Liquidity","Idiosyncratic"]
                   if c in decomp.columns]
fcolors = ["#"+FACTOR_HEX.get(c,"566573") for c in factor_cols_all]

# Chart 1: Stacked factor attribution
fig, ax = plt.subplots(figsize=(10, 4))
bottom = np.zeros(len(port_names))
for col, color in zip(factor_cols_all, fcolors):
    vals = np.array([float(decomp.loc[n,col]) for n in port_names])
    bars = ax.bar(port_names, vals, bottom=bottom, color=color, label=col, width=0.5)
    for bar, b, v in zip(bars, bottom, vals):
        if abs(v) > 4:
            ax.text(bar.get_x()+bar.get_width()/2, b+v/2, f"{v:.0f}%",
                    ha="center", va="center", fontsize=8, color="white", fontweight="bold")
    bottom += vals
ax.axhline(y=0, color="#1E2D3D", linewidth=0.8)
ax.set_ylabel("% of Total Portfolio Risk")
ax.set_title("Factor Risk Attribution by Construction Method", fontsize=11, fontweight="bold", color="#F0F4F8")
ax.legend(loc="upper right", fontsize=8, framealpha=0.3)
ax.set_ylim(-10, 115)
ax.grid(axis="y", alpha=0.4)
img = img_to_xl(fig)
img.width = 700; img.height = 300
ws2.add_image(img, "C3")

# Chart 2: ERP concentration bar
fig, ax = plt.subplots(figsize=(8, 3))
erp_vals = [float(decomp.loc[n,"Equity Premium"]) for n in port_names]
colors   = [("#B03A2E" if v>70 else "#B7770D" if v>55 else "#1E7B45") for v in erp_vals]
bars = ax.bar(port_names, erp_vals, color=colors, width=0.5)
for bar, v in zip(bars, erp_vals):
    ax.text(bar.get_x()+bar.get_width()/2, v+1, f"{v:.1f}%",
            ha="center", va="bottom", fontsize=9, fontweight="bold", color=bar.get_facecolor())
ax.axhline(y=50, color="#7A90A8", linewidth=1, linestyle="--", alpha=0.7, label="50% benchmark")
ax.set_ylabel("Equity Premium Share (%)")
ax.set_title("Equity Premium Concentration", fontsize=11, fontweight="bold", color="#F0F4F8")
ax.set_ylim(0, 105); ax.legend(fontsize=8, framealpha=0.3); ax.grid(axis="y", alpha=0.4)
img = img_to_xl(fig)
img.width = 560; img.height = 230
ws2.add_image(img, "C22")

# Chart 3: Return / Vol / Sharpe grouped bar
fig, axes = plt.subplots(1, 3, figsize=(12, 3.5))
metric_labels = ["Annualised Return (%)", "Annualised Volatility (%)", "Sharpe Ratio"]
metric_vals = [
    [metrics[n][0] for n in port_names],
    [metrics[n][1] for n in port_names],
    [metrics[n][2] for n in port_names],
]
for ax, label, vals in zip(axes, metric_labels, metric_vals):
    bars = ax.bar(port_names, vals, color=port_colors, width=0.55)
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x()+bar.get_width()/2, v+max(vals)*0.02,
                f"{v:.2f}" if "Sharpe" in label else f"{v:.1f}",
                ha="center", va="bottom", fontsize=8, color="white")
    ax.set_title(label, fontsize=9, fontweight="bold", color="#F0F4F8")
    ax.set_xticklabels(port_names, rotation=20, ha="right", fontsize=8)
    ax.grid(axis="y", alpha=0.4)
img = img_to_xl(fig)
img.width = 840; img.height = 270
ws2.add_image(img, "C38")

# Chart 4: Max drawdown and diversification ratio
fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 3.2))
mdd_vals = [metrics[n][4] for n in port_names]
bars = ax1.bar(port_names, mdd_vals, color=port_colors, width=0.5)
for bar, v in zip(bars, mdd_vals):
    ax1.text(bar.get_x()+bar.get_width()/2, v-0.5, f"{v:.1f}%",
             ha="center", va="top", fontsize=8, color="white")
ax1.set_title("Maximum Drawdown (%)", fontsize=10, fontweight="bold", color="#F0F4F8")
ax1.set_xticklabels(port_names, rotation=20, ha="right", fontsize=8)
ax1.grid(axis="y", alpha=0.4)

div_vals = []
for n, w in portfolios.items():
    wv   = w.reindex(cov.index).fillna(0).values
    if wv.sum() != 0:
        wv = wv / wv.sum()
    wvol = float(np.sum(wv * np.sqrt(np.diag(cov.values))))
    pvol = float(np.sqrt(wv @ cov.values @ wv))
    div_vals.append(wvol/pvol if pvol>0 else 1.0)
bars = ax2.bar(port_names, div_vals, color=port_colors, width=0.5)
for bar, v in zip(bars, div_vals):
    ax2.text(bar.get_x()+bar.get_width()/2, v+0.01, f"{v:.2f}x",
             ha="center", va="bottom", fontsize=8, color="white")
ax2.set_title("Diversification Ratio", fontsize=10, fontweight="bold", color="#F0F4F8")
ax2.set_xticklabels(port_names, rotation=20, ha="right", fontsize=8)
ax2.grid(axis="y", alpha=0.4)
plt.tight_layout()
img = img_to_xl(fig)
img.width = 700; img.height = 260
ws2.add_image(img, "C58")

# Chart 5: Cumulative return paths
fig, ax = plt.subplots(figsize=(11, 4))
for n, w in portfolios.items():
    _, _, _, _, _, _, _, pr = metrics[n]
    cum = (1 + pr).cumprod() * 100
    ax.plot(cum.index, cum.values, label=n, color="#"+PORT_HEX[n], linewidth=1.5)
if dm.recession is not None:
    rec_q = dm.recession.resample("QE").last().reindex(pr.index).fillna(0)
    in_rec = False; rec_start = None
    for date, val in rec_q.items():
        if val==1 and not in_rec: rec_start=date; in_rec=True
        elif val==0 and in_rec:
            ax.axvspan(rec_start, date, color="#333", alpha=0.25)
            in_rec=False
ax.set_title("Cumulative Growth of $100 — All Portfolios (2004 Q4 – 2024 Q4)", fontsize=11, fontweight="bold", color="#F0F4F8")
ax.set_ylabel("Portfolio Value ($)")
ax.legend(fontsize=9, framealpha=0.3)
ax.grid(alpha=0.4)
img = img_to_xl(fig)
img.width = 770; img.height = 310
ws2.add_image(img, "C76")

# ═══ SHEET 3: FACTOR MODEL ════════════════════════════════════════
ws3 = wb.create_sheet("Factor Model")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 28
for col in ["B","C","D","E","F","G","H","I","J"]:
    ws3.column_dimensions[col].width = 16

section_title(ws3, 1, 1, "Asset-Level Factor Model Results", 9)

# Beta matrix
header_row(ws3, 2, ["Asset Class","Equity Premium","Term Premium","Credit Spread","Inflation","Liquidity","Alpha","R-Squared","Liq. Proxy"],
           [28,16,16,16,16,16,14,14,20])
for i, asset in enumerate(result.betas.index):
    r2v   = float(result.r_squared.loc[asset]) if asset in result.r_squared.index else 0
    alpha = float(result.betas.loc[asset,"alpha"]) if "alpha" in result.betas.columns else 0
    beta_vals = [float(result.betas.loc[asset, f]) for f in result.betas.columns if f != "alpha"]
    from factors.factor_model import EQUITY_LIQUIDITY_ASSETS
    liq_proxy = "Pastor-Stambaugh" if asset in EQUITY_LIQUIDITY_ASSETS else "BAA-Treasury"
    bg = bg_alt[i % 2]
    data_row(ws3, 3+i,
             [dn(asset)] + beta_vals + [alpha, r2v, liq_proxy],
             bg=bg, fmt_map={2:"0.000",3:"0.000",4:"0.000",5:"0.000",6:"0.000",7:"0.0%",8:"0.0%"})
    ws3.cell(3+i, 1).alignment = aln("left")
    ws3.cell(3+i, 1).font = font("F0F4F8", 10)
    r2_cell = ws3.cell(3+i, 8)
    if r2v >= 0.7: r2_cell.font = Font(name="Arial", color=GREEN, size=10, bold=True)
    elif r2v >= 0.4: r2_cell.font = Font(name="Arial", color=AMBER, size=10, bold=True)
    else: r2_cell.font = Font(name="Arial", color=RED, size=10)

n_assets = len(result.betas.index)
ws3.row_dimensions[3+n_assets].height = 8

# Stress beta table
row_start = 4 + n_assets
section_title(ws3, row_start, 1, "Equity Premium Beta: Full-Sample OLS vs Stress (Q10 Quantile Regression)", 5)
header_row(ws3, row_start+1, ["Asset Class","OLS Beta","Stress Beta (Q10)","Uplift","Interpretation"],
           [28,16,20,16,30])
ols_erp = result.betas["equity_premium"]
q10_result2 = qm.results.get(0.10, None)
q10_erp  = q10_result2.betas["equity_premium"].reindex(ols_erp.index) if q10_result2 is not None else ols_erp.copy()
stress_rows = []
for asset in ols_erp.index:
    ols_b   = float(ols_erp.loc[asset])
    q10_b   = float(q10_erp.loc[asset]) if asset in q10_erp.index else ols_b
    uplift  = q10_b - ols_b
    interp  = ("Significant stress uplift" if uplift > 0.15 else
               "Moderate stress uplift" if uplift > 0.05 else
               "Stable across regimes")
    stress_rows.append((dn(asset), ols_b, q10_b, uplift, interp))
stress_rows.sort(key=lambda x: x[3], reverse=True)
for i, row in enumerate(stress_rows):
    bg = bg_alt[i % 2]
    data_row(ws3, row_start+2+i, list(row), bg=bg,
             fmt_map={2:"0.000",3:"0.000",4:"0.000"})
    ws3.cell(row_start+2+i, 1).alignment = aln("left")
    ws3.cell(row_start+2+i, 1).font = font("F0F4F8", 10)
    uplift_cell = ws3.cell(row_start+2+i, 4)
    if row[3] > 0.15: uplift_cell.font = Font(name="Arial", color=RED, size=10, bold=True)
    elif row[3] > 0.05: uplift_cell.font = Font(name="Arial", color=AMBER, size=10)

# ── Chart for stress beta (placed BEFORE private assets table) ────
fig, axes = plt.subplots(1, 2, figsize=(13, 4))
r2_sorted = result.r_squared.sort_values(ascending=True)
asset_colors = plt.cm.Blues(np.linspace(0.4, 0.9, len(r2_sorted)))
axes[0].barh([dn(a) for a in r2_sorted.index], r2_sorted.values, color=asset_colors)
axes[0].axvline(x=0.7, color="#B7770D", linestyle="--", linewidth=1, label="0.70 reference")
for i, v in enumerate(r2_sorted.values):
    axes[0].text(v+0.01, i, f"{v:.3f}", va="center", fontsize=8, color="#D0D8E4")
axes[0].set_title("Adjusted R-Squared by Asset Class", fontsize=10, fontweight="bold", color="#F0F4F8")
axes[0].set_xlim(0, 1.15); axes[0].legend(fontsize=8, framealpha=0.3)
axes[0].grid(axis="x", alpha=0.4)

sorted_idx = sorted(ols_erp.index, key=lambda a: float(ols_erp.loc[a]))
x = np.arange(len(sorted_idx))
axes[1].barh(x - 0.2, [float(ols_erp.loc[a]) for a in sorted_idx],
             height=0.35, label="OLS (Full Sample)", color="#2980B9")
axes[1].barh(x + 0.2, [float(q10_erp.loc[a]) if a in q10_erp.index else float(ols_erp.loc[a])
                        for a in sorted_idx],
             height=0.35, label="Stress Q10", color="#B03A2E", alpha=0.85)
axes[1].set_yticks(x); axes[1].set_yticklabels([dn(a) for a in sorted_idx], fontsize=8)
axes[1].set_title("Equity Beta: Normal vs Stress", fontsize=10, fontweight="bold", color="#F0F4F8")
axes[1].legend(fontsize=8, framealpha=0.3); axes[1].grid(axis="x", alpha=0.4)
plt.tight_layout()
img = img_to_xl(fig)
img.width = 910; img.height = 310
# Chart goes right below the stress table, with a 2-row gap
chart_row_stress = row_start + 2 + len(stress_rows) + 2
ws3.add_image(img, f"A{chart_row_stress}")

# ── Private Assets Alpha vs Beta Decomposition ───────────────────
# Placed BELOW the stress chart (chart is ~24 rows tall at 310px / ~13px per row)
CHART_ROW_HEIGHT = 24          # approximate rows consumed by a 310px chart at 13px/row
row_private = chart_row_stress + CHART_ROW_HEIGHT + 1

PRIVATE_ASSETS = ["private_equity","private_credit","infrastructure","real_assets"]
private_in_model = [a for a in PRIVATE_ASSETS if a in result.betas.index]

# ── FIX: Correctly pre-compute annualised mean factor returns (decimal) ──
# mean_quarterly_factor_rets is already computed at the top as a dict {factor: float (quarterly decimal)}
# Annualised = quarterly_mean * 4  →  stays as a decimal (do NOT multiply by 100 here)
ann_factor_rets = {
    f: mean_quarterly_factor_rets[f] * 4
    for f in factor_col_keys
    if f in mean_quarterly_factor_rets
}

section_title(ws3, row_private, 1,
    "Alpha vs. Beta Decomposition — Private Assets (Post-Unsmoothing)", 8)
header_row(ws3, row_private + 1,
    ["Asset Class",
     "OLS Alpha (qtly)",      # col 2: raw quarterly intercept from OLS (decimal)
     "Multi-Factor Beta Ret", # col 3: Σ(β_f × mean_ann_factor_ret) — total factor-explained return (decimal)
     "Equity Beta Contrib",   # col 4: equity_beta × mean_ann_equity_factor_ret (decimal)
     "True Alpha (ex-all β)", # col 5: annualised alpha net of ALL factor betas (decimal)
     "Alpha vs Equity-Only",  # col 6: annualised alpha net of equity beta alone (decimal)
     "Interpretation",
     "R-Squared"],
    [28, 20, 22, 22, 22, 22, 42, 14])

for i, asset in enumerate(private_in_model):
    # ── Raw OLS intercept (quarterly, decimal) ───────────────────
    raw_alpha_qtly = float(result.betas.loc[asset, "alpha"]) if "alpha" in result.betas.columns else 0.0

    # ── Annualised OLS intercept (decimal) ───────────────────────
    # Quarterly alpha * 4 gives annualised (geometric approximation is fine for small α)
    raw_alpha_ann = raw_alpha_qtly * 4

    # ── Multi-factor beta return contribution (annualised, decimal) ──
    # = Σ_f ( β_{asset,f} × mean_ann_factor_return_f )
    # This is the portion of expected return explained by ALL systematic factor betas.
    multi_factor_beta_ret = 0.0
    for f in factor_col_keys:
        if f in result.betas.columns and f in ann_factor_rets:
            beta_f = float(result.betas.loc[asset, f])
            multi_factor_beta_ret += beta_f * ann_factor_rets[f]

    # ── Equity beta contribution alone (annualised, decimal) ─────
    eq_beta = float(result.betas.loc[asset, "equity_premium"]) if "equity_premium" in result.betas.columns else 0.0
    eq_ann_ret = ann_factor_rets.get("equity_premium", 0.0)
    equity_beta_contrib = eq_beta * eq_ann_ret

    # ── True alpha = annualised OLS intercept − full multi-factor beta return ──
    # This isolates the return unexplained by ANY systematic factor exposure.
    # NOTE: In an OLS factor model the intercept already IS the residual from factor
    # regression, so raw_alpha_ann is the "true alpha" relative to the factor model.
    # multi_factor_beta_ret is the factor-model fitted return (excl. intercept).
    # Combining:  expected_return ≈ raw_alpha_ann + multi_factor_beta_ret
    # True alpha (net of all betas) = raw_alpha_ann  ← this IS the OLS intercept
    # Alpha net of equity beta only = raw_alpha_ann + (multi_factor_beta_ret - equity_beta_contrib)
    #   = raw_alpha_ann + non_equity_factor_contributions
    true_alpha_ex_all  = raw_alpha_ann          # net of ALL factor betas (= OLS intercept annualised)
    alpha_ex_equity    = raw_alpha_ann + (multi_factor_beta_ret - equity_beta_contrib)
    # alpha_ex_equity tells us: "if you strip equity beta but keep other factor betas,
    # how much residual return does the asset generate?"

    r2v = float(result.r_squared.loc[asset]) if asset in result.r_squared.index else 0.0

    # ── Interpretation logic ─────────────────────────────────────
    if r2v < 0.3:
        interp = "Low R² — stale pricing / smoothed returns likely inflate apparent alpha; treat with caution"
    elif abs(true_alpha_ex_all) < 0.005:   # < 0.5% annualised
        interp = "Near-zero true alpha — return is primarily leveraged systematic beta; fee load erodes value"
    elif true_alpha_ex_all > 0.01:          # > 1% annualised
        interp = "Positive true alpha (ex-all β) — potential illiquidity/manager premium above factor exposure"
    elif true_alpha_ex_all < -0.005:
        interp = "Negative true alpha — systematic beta alone does not justify the allocation; re-evaluate fees"
    else:
        interp = "Marginal alpha — monitor whether fee load and illiquidity premium net to positive value-add"

    bg = bg_alt[i % 2]
    # All return figures stored as decimals → formatted as 0.00% in Excel
    data_row(ws3, row_private + 2 + i,
             [dn(asset),
              raw_alpha_qtly,           # col 2: quarterly decimal
              multi_factor_beta_ret,    # col 3: annualised decimal
              equity_beta_contrib,      # col 4: annualised decimal
              true_alpha_ex_all,        # col 5: annualised decimal
              alpha_ex_equity,          # col 6: annualised decimal
              interp,
              r2v],
             bg=bg,
             fmt_map={
                 2: "0.000",    # quarterly intercept: show 3dp as a small number
                 3: "0.00%",    # multi-factor beta ret (annualised %)
                 4: "0.00%",    # equity beta contrib (annualised %)
                 5: "0.00%",    # true alpha ex-all (annualised %)
                 6: "0.00%",    # alpha ex-equity (annualised %)
                 8: "0.0%",     # R-squared
             })

    ws3.cell(row_private + 2 + i, 1).alignment = aln("left")
    ws3.cell(row_private + 2 + i, 1).font = font("F0F4F8", 10)
    ws3.cell(row_private + 2 + i, 7).alignment = aln("left", wrap=True)

    # Colour true alpha cell (col 5)
    alpha_cell = ws3.cell(row_private + 2 + i, 5)
    if true_alpha_ex_all > 0.01:
        alpha_cell.font = Font(name="Arial", color=GREEN, size=10, bold=True)
    elif true_alpha_ex_all < -0.005:
        alpha_cell.font = Font(name="Arial", color=RED, size=10, bold=True)
    else:
        alpha_cell.font = Font(name="Arial", color=AMBER, size=10)

    # Colour R-squared cell (col 8)
    r2_cell = ws3.cell(row_private + 2 + i, 8)
    if r2v >= 0.7:
        r2_cell.font = Font(name="Arial", color=GREEN, size=10, bold=True)
    elif r2v >= 0.4:
        r2_cell.font = Font(name="Arial", color=AMBER, size=10)
    else:
        r2_cell.font = Font(name="Arial", color=RED, size=10)

# Footnote explaining the decomposition
row_fn = row_private + 2 + len(private_in_model) + 1
ws3.merge_cells(f"A{row_fn}:H{row_fn}")
c = ws3[f"A{row_fn}"]
c.value = (
    "Decomposition: OLS regression → Asset Return = α + Σ(β_f × Factor_f) + ε.  "
    "Multi-Factor Beta Ret = Σ(β_f × mean_ann_factor_ret) across all 5 factors.  "
    "True Alpha (ex-all β) = annualised OLS intercept — the return unexplained by any systematic factor.  "
    "Alpha vs Equity-Only strips only equity beta, retaining other factor contributions.  "
    "Low R² assets use smoothed/stale prices; beta estimates are understated → alpha overstated."
)
c.fill = fill(NAVY); c.font = font(MUTED, 9)
c.alignment = aln("left", wrap=True)
ws3.row_dimensions[row_fn].height = 44

# ═══ SHEET: LOOK-THROUGH FACTOR EXPOSURE ═════════════════════════
ws_lt = wb.create_sheet("Look-Through Exposure")
ws_lt.sheet_view.showGridLines = False
ws_lt.column_dimensions["A"].width = 28
for col in ["B","C","D","E","F","G"]: ws_lt.column_dimensions[col].width = 20

section_title(ws_lt, 1, 1,
    "Look-Through Net Factor Exposure — Portfolio Weight × Asset Beta", 7)

factor_cols_lt = ["equity_premium","term_premium","credit_spread","inflation","liquidity"]
factor_display_lt = ["Equity Premium","Term Premium","Credit Spread","Inflation","Liquidity"]

header_row(ws_lt, 2,
    ["Portfolio / Asset"] + factor_display_lt + ["Eff. Equity Weight*"],
    [28]+[18]*5+[22])

for port_name, w in portfolios.items():
    wv = w.reindex(result.betas.index).fillna(0)
    wv_norm = wv / wv.sum()

    port_row = [f"▶ {port_name}"]
    for f_col in factor_cols_lt:
        if f_col in result.betas.columns:
            net_exp = float((wv_norm * result.betas[f_col]).sum())
        else:
            net_exp = 0.0
        port_row.append(net_exp)
    if "equity_premium" in result.betas.columns:
        eff_eq = float((wv_norm * result.betas["equity_premium"]).sum())
    else:
        eff_eq = 0.0
    port_row.append(eff_eq)

    r = ws_lt.max_row + 1
    data_row(ws_lt, r, port_row, bg=BLUE,
             fmt_map={j+2:"0.000" for j in range(6)})
    ws_lt.cell(r, 1).font = font("F0F4F8", 11, True)
    ws_lt.cell(r, 1).alignment = aln("left")

    for asset in wv_norm.index:
        w_i = float(wv_norm[asset])
        if w_i < 0.001: continue
        asset_row = [f"    {dn(asset)}"]
        for f_col in factor_cols_lt:
            if f_col in result.betas.columns and asset in result.betas.index:
                contrib = w_i * float(result.betas.loc[asset, f_col])
            else:
                contrib = 0.0
            asset_row.append(contrib)
        if "equity_premium" in result.betas.columns and asset in result.betas.index:
            eff_eq_i = w_i * float(result.betas.loc[asset, "equity_premium"])
        else:
            eff_eq_i = 0.0
        asset_row.append(eff_eq_i)
        r2 = ws_lt.max_row + 1
        data_row(ws_lt, r2, asset_row, bg=PANEL,
                 fmt_map={j+2:"0.000" for j in range(6)})
        ws_lt.cell(r2, 1).alignment = aln("left")
        ws_lt.cell(r2, 1).font = font("D0D8E4", 9)

# Footnote
r_note = ws_lt.max_row + 2
ws_lt.merge_cells(f"A{r_note}:G{r_note}")
c = ws_lt[f"A{r_note}"]
c.value = ("* Effective Equity Weight = Σ(w_i × equity_beta_i). "
           "Example: 10% Private Equity with beta 0.85 contributes 0.085 to net equity exposure. "
           "A value >1.0 implies leveraged equity risk.")
c.fill = fill(NAVY); c.font = font(MUTED, 9)
c.alignment = aln("left", wrap=True)

# Grouped bar chart of net factor exposures
fig, ax = plt.subplots(figsize=(8, 5))
x = np.arange(len(factor_display_lt))
bar_w = 0.15
for j, (port_name, w) in enumerate(portfolios.items()):
    wv = w.reindex(result.betas.index).fillna(0)
    wv = wv / wv.sum()
    exps = []
    for f_col in factor_cols_lt:
        if f_col in result.betas.columns:
            exps.append(float((wv * result.betas[f_col]).sum()))
        else:
            exps.append(0.0)
    offset = (j - 2) * bar_w
    ax.bar(x + offset, exps, bar_w, label=port_name,
           color="#"+PORT_HEX[port_name], alpha=0.85)
ax.axhline(y=0, color="#7A90A8", linewidth=0.8)
ax.set_xticks(x); ax.set_xticklabels(factor_display_lt, fontsize=10)
ax.set_ylabel("Net Factor Exposure (β-weighted)")
ax.set_title("Look-Through Net Factor Exposure by Portfolio",
             fontsize=11, fontweight="bold", color="#F0F4F8")
ax.legend(fontsize=8, framealpha=0.3)
ax.grid(axis="y", alpha=0.4)
img = img_to_xl(fig)
img.width = 700; img.height = 300
ws_lt.add_image(img, f"A{r_note+3}")


# ═══ SHEET 4: STRESS TESTING ══════════════════════════════════════
ws4 = wb.create_sheet("Stress Testing")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 34
for col in ["B","C","D","E","F"]: ws4.column_dimensions[col].width = 20

section_title(ws4, 1, 1, "Portfolio Stress Testing — Defined Periods", 6)
header_row(ws4, 2,
    ["Portfolio"] + list(stress_periods.keys()) + ["Recovery: 2009 Q2 – 2010 Q4"],
    [34]+[20]*len(stress_periods)+[28])

for i, (name, w) in enumerate(portfolios.items()):
    _, _, _, _, _, _, _, pr = metrics[name]
    row_data = [name]
    for label, (start, end) in stress_periods.items():
        row_data.append(period_return(pr, start, end)/100)
    row_data.append(period_return(pr, "2009-04-01", "2010-12-31")/100)
    bg = bg_alt[i % 2]
    data_row(ws4, 3+i, row_data, bg=bg,
             fmt_map={j+2:"0.0%" for j in range(len(stress_periods)+1)})
    ws4.cell(3+i, 1).alignment = aln("left")
    ws4.cell(3+i, 1).font = font("F0F4F8", 10, True)
    for j in range(2, 2+len(stress_periods)):
        cell = ws4.cell(3+i, j)
        if cell.value is not None and isinstance(cell.value, float):
            if cell.value < -0.08: cell.font = Font(name="Arial", color=RED, size=10, bold=True)
            elif cell.value < -0.03: cell.font = Font(name="Arial", color=AMBER, size=10)
            elif cell.value > 0: cell.font = Font(name="Arial", color=GREEN, size=10)

ws4.row_dimensions[3+len(portfolios)].height = 8

row2 = 4 + len(portfolios) + 1
section_title(ws4, row2, 1, "Asset Class Returns During Stress Periods", 5)
header_row(ws4, row2+1,
    ["Asset Class"] + list(stress_periods.keys()) + ["Recovery: 2009 Q2 – 2010 Q4"],
    [34]+[20]*len(stress_periods)+[28])

asset_rets = dm.asset_returns_t1_complete
for i, asset in enumerate(asset_rets.columns):
    row_data = [dn(asset)]
    ar = asset_rets[asset].dropna()
    for label, (start, end) in stress_periods.items():
        row_data.append(period_return(ar, start, end)/100)
    row_data.append(period_return(ar, "2009-04-01", "2010-12-31")/100)
    bg = bg_alt[i % 2]
    data_row(ws4, row2+2+i, row_data, bg=bg,
             fmt_map={j+2:"0.0%" for j in range(len(stress_periods)+1)})
    ws4.cell(row2+2+i, 1).alignment = aln("left")
    ws4.cell(row2+2+i, 1).font = font("F0F4F8", 10)
    for j in range(2, 2+len(stress_periods)):
        cell = ws4.cell(row2+2+i, j)
        if cell.value is not None and isinstance(cell.value, float):
            if cell.value < -0.12: cell.font = Font(name="Arial", color=RED, size=10, bold=True)
            elif cell.value < -0.04: cell.font = Font(name="Arial", color=AMBER, size=10)
            elif cell.value > 0.05: cell.font = Font(name="Arial", color=GREEN, size=10)

# Stress chart
fig, ax = plt.subplots(figsize=(11, 4))
x = np.arange(len(port_names))
period_keys = list(stress_periods.keys())
bar_width = 0.18
for j, (pname, (start, end)) in enumerate(stress_periods.items()):
    vals = []
    for n, w in portfolios.items():
        _, _, _, _, _, _, _, pr = metrics[n]
        vals.append(period_return(pr, start, end))
    offset = (j - 1) * bar_width
    bars = ax.bar(x + offset, vals, bar_width,
                  label=pname, alpha=0.85)
    for bar, v in zip(bars, vals):
        if not np.isnan(v):
            ax.text(bar.get_x()+bar.get_width()/2, v + (0.3 if v >= 0 else -0.5),
                    f"{v:.1f}%", ha="center", va="bottom" if v >= 0 else "top",
                    fontsize=7, color="white")
ax.axhline(y=0, color="#7A90A8", linewidth=0.8)
ax.set_xticks(x); ax.set_xticklabels(port_names, fontsize=9)
ax.set_ylabel("Total Return (%)")
ax.set_title("Portfolio Returns During Stress Periods", fontsize=11, fontweight="bold", color="#F0F4F8")
ax.legend(fontsize=8, framealpha=0.3)
ax.grid(axis="y", alpha=0.4)
img = img_to_xl(fig)
img.width = 770; img.height = 310
ws4.add_image(img, f"A{row2+2+len(asset_rets.columns)+2}")


# ═══ SHEET 5: RISK DECOMPOSITION ══════════════════════════════════
ws5 = wb.create_sheet("Risk Decomposition")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions["A"].width = 28
for col in ["B","C","D","E","F","G"]: ws5.column_dimensions[col].width = 18

section_title(ws5, 1, 1, "Asset Marginal Risk Contribution by Portfolio", 6)

for port_idx, (name, w) in enumerate(portfolios.items()):
    row_off = 2 + port_idx * (len(assets) + 3)
    section_title(ws5, row_off, 1, f"{name} — Marginal and Total Risk Contribution", 6)
    header_row(ws5, row_off+1,
        ["Asset Class","Weight","Marginal Risk Contrib.","Total Risk Contrib.","% of Portfolio Risk","Factor Risk Share"],
        [28,14,22,22,22,20])
    wv  = w.reindex(cov.index).fillna(0).values
    if wv.sum() != 0:
        wv = wv / wv.sum()
    cov_v = cov.values
    port_vol = float(np.sqrt(wv @ cov_v @ wv))
    mrc = (cov_v @ wv) / port_vol if port_vol > 0 else wv*0
    trc = wv * mrc
    for i, asset in enumerate(cov.index):
        wt    = float(w.get(asset, 0))
        mrc_i = float(mrc[i])
        trc_i = float(trc[i])
        pct_i = trc_i / port_vol * 100 if port_vol > 0 else 0
        # Use factor-only betas for factor risk share
        if asset in result.betas.index:
            b_i = result.betas.loc[asset, factor_col_keys].values.astype(float)
        else:
            b_i = np.zeros(len(factor_col_keys))
        syst_i  = float(b_i @ factor_cov.values @ b_i)
        tot_i   = float(cov_v[i,i])
        fshr    = syst_i/tot_i if tot_i > 0 else 0
        bg = bg_alt[i % 2]
        data_row(ws5, row_off+2+i,
                 [dn(asset), wt, mrc_i, trc_i, pct_i/100, fshr],
                 bg=bg, fmt_map={2:"0.0%",3:"0.000",4:"0.000",5:"0.0%",6:"0.0%"})
        ws5.cell(row_off+2+i, 1).alignment = aln("left")
        ws5.cell(row_off+2+i, 1).font = font("F0F4F8", 10)

# MRC chart
fig, axes = plt.subplots(2, 3, figsize=(14, 8))
axes = axes.flatten()
for idx, (name, w) in enumerate(portfolios.items()):
    ax = axes[idx]
    wv = w.reindex(cov.index).fillna(0).values
    if wv.sum() != 0:
        wv = wv / wv.sum()
    cov_v = cov.values
    port_vol = float(np.sqrt(wv @ cov_v @ wv))
    trc = wv * ((cov_v @ wv) / port_vol) if port_vol > 0 else wv
    pct = trc / port_vol * 100
    sorted_idx = np.argsort(pct)[::-1]
    labels_s   = [dn(cov.index[i]) for i in sorted_idx]
    vals_s     = [pct[i] for i in sorted_idx]
    c_list     = ["#B03A2E" if v > 15 else "#2980B9" if v > 5 else "#1E7B45" for v in vals_s]
    ax.barh(labels_s, vals_s, color=c_list)
    ax.set_title(name, fontsize=9, fontweight="bold", color="#F0F4F8")
    ax.set_xlabel("% of Portfolio Risk", fontsize=8)
    ax.tick_params(axis="y", labelsize=7)
    ax.grid(axis="x", alpha=0.4)
if len(portfolios) < 6:
    for idx in range(len(portfolios), 6): axes[idx].set_visible(False)
plt.suptitle("Asset Marginal Risk Contribution by Portfolio", fontsize=11,
             fontweight="bold", color="#F0F4F8", y=1.01)
plt.tight_layout()
img = img_to_xl(fig)
img.width = 980; img.height = 600
row_chart = 2 + len(portfolios) * (len(assets) + 3) + 1
ws5.add_image(img, f"A{row_chart}")

# ═══ FACTOR PERFORMANCE ATTRIBUTION ══════════════════════════════
ws_fa = wb.create_sheet("Factor Performance")
ws_fa.sheet_view.showGridLines = False
ws_fa.column_dimensions["A"].width = 28
for col in ["B","C","D","E","F","G"]: ws_fa.column_dimensions[col].width = 18

section_title(ws_fa, 1, 1, "Factor Performance Attribution — Brinson-Style Decomposition", 7)

F_ann = dm.factor_returns_t1.mean() * 4 * 100   # annualised %

header_row(ws_fa, 2,
    ["Factor","Realised Ann. Return (%)","EW Attribution","60/40 Attribution",
     "MVO Attribution","Risk Parity Attrib.","Enhanced HRP Attrib."],
    [28]+[20]*6)

factor_display_map = {
    "equity_premium": "Equity Premium",
    "term_premium":   "Term Premium",
    "credit_spread":  "Credit Spread",
    "inflation":      "Inflation",
    "liquidity":      "Liquidity",
}

for i, (f_col, f_label) in enumerate(factor_display_map.items()):
    if f_col not in F_ann.index: continue
    f_ret = float(F_ann[f_col])
    row_data = [f_label, f_ret/100]
    for name, w in portfolios.items():
        wv = w.reindex(result.betas.index).fillna(0)
        wv = wv / wv.sum()
        if f_col in result.betas.columns:
            port_beta = float((wv * result.betas[f_col]).sum())
        else:
            port_beta = 0.0
        attrib = port_beta * f_ret / 100   # beta × factor return, expressed as decimal
        row_data.append(attrib)
    bg = bg_alt[i % 2]
    fmt = {2:"0.0%"}
    fmt.update({j+3:"0.0%" for j in range(len(portfolios))})
    data_row(ws_fa, 3+i, row_data, bg=bg, fmt_map=fmt)
    ws_fa.cell(3+i, 1).alignment = aln("left")
    ws_fa.cell(3+i, 1).font = font("F0F4F8", 10)
    ret_cell = ws_fa.cell(3+i, 2)
    if f_ret > 2: ret_cell.font = Font(name="Arial", color=GREEN, size=10, bold=True)
    elif f_ret < -1: ret_cell.font = Font(name="Arial", color=RED, size=10)

row_alpha = 3 + len(factor_display_map)
ws_fa.cell(row_alpha, 1, "Alpha (Unexplained)").fill = fill(PANEL)
ws_fa.cell(row_alpha, 1).font = font("F0F4F8", 10, True)
ws_fa.cell(row_alpha, 1).alignment = aln("left")
ws_fa.cell(row_alpha, 1).border = border_thin()

# Attribution bar chart
fig, ax = plt.subplots(figsize=(10, 4))
factor_labels_short = list(factor_display_map.values())
x = np.arange(len(factor_labels_short))
bar_w = 0.15
for j, (name, w) in enumerate(portfolios.items()):
    attribs = []
    wv = w.reindex(result.betas.index).fillna(0)
    wv = wv / wv.sum()
    for f_col in factor_display_map:
        if f_col not in F_ann.index or f_col not in result.betas.columns:
            attribs.append(0)
            continue
        port_beta = float((wv * result.betas[f_col]).sum())
        attribs.append(port_beta * float(F_ann[f_col]))
    offset = (j - 2) * bar_w
    ax.bar(x + offset, attribs, bar_w, label=name,
           color="#"+PORT_HEX[name], alpha=0.85)
ax.axhline(y=0, color="#7A90A8", linewidth=0.8)
ax.set_xticks(x); ax.set_xticklabels(factor_labels_short, rotation=15, ha="right", fontsize=9)
ax.set_ylabel("Return Attribution (% ann.)")
ax.set_title("Factor Performance Attribution — Beta × Realised Factor Return",
             fontsize=11, fontweight="bold", color="#F0F4F8")
ax.legend(fontsize=8, framealpha=0.3)
ax.grid(axis="y", alpha=0.4)
img = img_to_xl(fig)
img.width = 700; img.height = 300
ws_fa.add_image(img, f"A{row_alpha+3}")

# ═══ SHEET 6: COVARIANCE ══════════════════════════════════════════
ws6 = wb.create_sheet("Covariance")
ws6.sheet_view.showGridLines = False
asset_labels = [dn(a) for a in cov.index]
n = len(asset_labels)

ws6.column_dimensions["A"].width = 26
for i in range(2, n+2):
    ws6.column_dimensions[get_column_letter(i)].width = 12

section_title(ws6, 1, 1, "POET Correlation Matrix", n+1)
for j, lab in enumerate(asset_labels, 2):
    cell = ws6.cell(2, j, lab)
    cell.fill = fill(NAVY); cell.font = font("F0F4F8", 8, True)
    cell.alignment = aln("center", wrap=True); cell.border = border_thin()
ws6.row_dimensions[2].height = 56

std  = np.sqrt(np.diag(cov.values))
corr = cov.values / np.outer(std, std)

for i, asset_row in enumerate(cov.index):
    cell = ws6.cell(3+i, 1, dn(asset_row))
    cell.fill = fill(NAVY); cell.font = font("F0F4F8", 9, True)
    cell.alignment = aln("left"); cell.border = border_thin()
    for j in range(n):
        val  = float(corr[i,j])
        cell = ws6.cell(3+i, 2+j, round(val, 3))
        if i == j:
            bg_c = "1E7B45"
        elif val > 0.7:
            bg_c = "6B1A1A"
        elif val > 0.4:
            bg_c = "8B2E1A"
        elif val < -0.2:
            bg_c = "1A3A6B"
        else:
            bg_c = PANEL
        cell.fill = fill(bg_c)
        cell.font = font("F0F4F8", 9)
        cell.alignment = aln("center")
        cell.border = border_thin()
        cell.number_format = "0.000"

ws6.row_dimensions[3+n].height = 8

row_pairs = 4 + n
section_title(ws6, row_pairs, 1, "Notable Correlation Pairs", 5)
header_row(ws6, row_pairs+1, ["Asset A","Asset B","Correlation","Type","Implication"],
           [26,26,14,20,40])
pairs = []
for i in range(n):
    for j in range(i+1, n):
        pairs.append((cov.index[i], cov.index[j], float(corr[i,j])))
pairs.sort(key=lambda x: abs(x[2]), reverse=True)
for k, (a, b, c_val) in enumerate(pairs[:15]):
    corr_type = "Strong Positive" if c_val > 0.6 else "Moderate Positive" if c_val > 0.3 else "Negative" if c_val < -0.1 else "Low"
    if c_val > 0.6: impl = "Low diversification benefit between these assets"
    elif c_val < -0.1: impl = "Hedging relationship — diversification benefit present"
    else: impl = "Moderate diversification benefit"
    bg = bg_alt[k % 2]
    data_row(ws6, row_pairs+2+k, [dn(a), dn(b), c_val, corr_type, impl],
             bg=bg, fmt_map={3:"0.000"})
    ws6.cell(row_pairs+2+k, 1).alignment = aln("left")
    ws6.cell(row_pairs+2+k, 2).alignment = aln("left")
    ws6.cell(row_pairs+2+k, 4).alignment = aln("left")
    ws6.cell(row_pairs+2+k, 5).alignment = aln("left")
    val_cell = ws6.cell(row_pairs+2+k, 3)
    if c_val > 0.6: val_cell.font = Font(name="Arial", color=RED, size=10, bold=True)
    elif c_val < -0.1: val_cell.font = Font(name="Arial", color=GREEN, size=10)

fig, ax = plt.subplots(figsize=(10, 8))
im = ax.imshow(corr, cmap="RdBu_r", vmin=-1, vmax=1, aspect="auto")
ax.set_xticks(range(n)); ax.set_xticklabels(asset_labels, rotation=45, ha="right", fontsize=8)
ax.set_yticks(range(n)); ax.set_yticklabels(asset_labels, fontsize=8)
for i in range(n):
    for j in range(n):
        ax.text(j, i, f"{corr[i,j]:.2f}", ha="center", va="center",
                fontsize=7, color="white" if abs(corr[i,j]) > 0.4 else "#D0D8E4")
plt.colorbar(im, ax=ax, fraction=0.04, label="Correlation")
ax.set_title("POET Correlation Matrix — All 13 Asset Classes", fontsize=11,
             fontweight="bold", color="#F0F4F8")
plt.tight_layout()
img = img_to_xl(fig)
img.width = 700; img.height = 580
ws6.add_image(img, f"G{row_pairs+2}")

# ═══ SHEET 7: PORTFOLIO WEIGHTS ═══════════════════════════════════
ws7 = wb.create_sheet("Portfolio Weights")
ws7.sheet_view.showGridLines = False
ws7.column_dimensions["A"].width = 28
for col in ["B","C","D","E","F"]: ws7.column_dimensions[col].width = 18

section_title(ws7, 1, 1, "Portfolio Weights — All Construction Methods (%)", 6)
header_row(ws7, 2, ["Asset Class"] + list(portfolios.keys()), [28]+[18]*len(portfolios))
for i, asset in enumerate(cov.index):
    row_data = [dn(asset)]
    for name, w in portfolios.items():
        row_data.append(float(w.get(asset, 0)))
    bg = bg_alt[i % 2]
    data_row(ws7, 3+i, row_data, bg=bg, fmt_map={j+2:"0.0%" for j in range(len(portfolios))})
    ws7.cell(3+i, 1).alignment = aln("left")
    ws7.cell(3+i, 1).font = font("F0F4F8", 10)
    for j, (name, _) in enumerate(portfolios.items()):
        cell = ws7.cell(3+i, 2+j)
        if cell.value and isinstance(cell.value, float) and cell.value >= 0.10:
            cell.font = Font(name="Arial", color="F0F4F8", size=10, bold=True)

fig, axes = plt.subplots(1, len(portfolios), figsize=(16, 5))
for idx, (name, w) in enumerate(portfolios.items()):
    ax = axes[idx]
    w_sorted = w.reindex(cov.index).fillna(0).sort_values(ascending=False)
    w_sorted = w_sorted[w_sorted > 0.005]
    labels_s = [dn(a) for a in w_sorted.index]
    vals_s   = w_sorted.values * 100
    bars = ax.barh(labels_s[::-1], vals_s[::-1],
                   color="#"+PORT_HEX[name], alpha=0.85)
    for bar, v in zip(bars, vals_s[::-1]):
        ax.text(v+0.3, bar.get_y()+bar.get_height()/2,
                f"{v:.1f}%", va="center", fontsize=7, color="#D0D8E4")
    ax.set_title(name, fontsize=9, fontweight="bold", color="#F0F4F8")
    ax.set_xlabel("Weight (%)", fontsize=8)
    ax.tick_params(axis="y", labelsize=7)
    ax.grid(axis="x", alpha=0.4)
plt.suptitle("Portfolio Weights by Construction Method", fontsize=11,
             fontweight="bold", color="#F0F4F8", y=1.02)
plt.tight_layout()
img = img_to_xl(fig)
img.width = 1120; img.height = 400
ws7.add_image(img, f"A{4+len(cov.index)+1}")

# ═══ SHEET 8: CLUSTER ANALYSIS ════════════════════════════════════
print("Building cluster analysis sheet...")
from scipy.cluster.hierarchy import linkage, fcluster, dendrogram
from scipy.spatial.distance import pdist, squareform
from portfolio.hrp import FACTOR_NAMES

ws8 = wb.create_sheet("Cluster Analysis")
ws8.sheet_view.showGridLines = False

col_widths = {"A":3,"B":28,"C":28,"D":28,"E":28,"F":28,"G":28,"H":28,"I":28,"J":28}
for col, w in col_widths.items():
    ws8.column_dimensions[col].width = w

# Compute clusters
B_cl = result.betas.reindex(assets)[FACTOR_NAMES].values.astype(float)
mu_b = B_cl.mean(axis=0); sig_b = B_cl.std(axis=0); sig_b[sig_b==0] = 1.0
B_z  = (B_cl - mu_b) / sig_b
link = linkage(B_z, method="ward")
labels_5 = fcluster(link, t=5, criterion="maxclust")

FACTOR_DISPLAY = ["Equity Premium","Term Premium","Credit Spread","Inflation","Liquidity"]
CLUSTER_META = {
    1: ("Yield-Seeking Diversifiers", "B03A2E",
        "Low equity beta complex. Credit and alternative assets sharing below-average equity exposure."),
    2: ("Pure Rate Duration",         "2E6DA4",
        "Positive term premium, negative equity beta. Sole rate hedge in universe."),
    3: ("Inflation Hedge",            "B7770D",
        "Dominant inflation beta. Real commodity dynamics orthogonal to equity and credit."),
    4: ("Domestic Equity Beta",       "1E7B45",
        "High positive equity loading across all three size segments. Essentially same systematic risk."),
    5: ("Credit-Sensitive Equity",    "6B3FA0",
        "High credit spread sensitivity, strongly negative liquidity beta. Seizes during credit stress."),
}

CLUSTER_DEFS = {}
for cid in range(1, 6):
    idx_list = [i for i in range(len(assets)) if labels_5[i]==cid]
    members  = [dn(assets[i]) for i in idx_list]
    centroid = B_z[idx_list].mean(axis=0) if idx_list else np.zeros(5)
    dom_idx  = int(np.argmax(np.abs(centroid)))
    CLUSTER_DEFS[cid] = {"members": members, "centroid": centroid,
                          "dominant": FACTOR_DISPLAY[dom_idx], "n": len(members)}

ws8.merge_cells("B1:J1")
c = ws8["B1"]
c.value = "FACTOR-LOADING CLUSTER ANALYSIS — ENHANCED HRP"
c.fill = fill(NAVY); c.font = font("F0F4F8", 14, True)
c.alignment = aln("center")
ws8.row_dimensions[1].height = 30

ws8.merge_cells("B2:J2")
c = ws8["B2"]
c.value = ("Ward linkage on z-scored factor betas  |  5-Factor Model  |  13 Asset Classes  |  "
           "Euclidean distance in standardised beta space  |  2004 Q4 to 2024 Q4")
c.fill = fill(BLUE); c.font = font(MUTED, 10)
c.alignment = aln("center")
ws8.row_dimensions[2].height = 18
ws8.row_dimensions[3].height = 10

ws8.merge_cells("B4:J4")
c = ws8["B4"]
c.value = "5-CLUSTER SOLUTION — ECONOMIC INTERPRETATION"
c.fill = fill(BLUE); c.font = font("F0F4F8", 11, True)
c.alignment = aln("left")
ws8.row_dimensions[4].height = 22

card_col_starts = [2, 4, 6, 8, 10]
from openpyxl.utils import get_column_letter

for cid in range(1, 6):
    meta  = CLUSTER_META[cid]
    defn  = CLUSTER_DEFS[cid]
    label, hex_clr, desc = meta
    col1 = card_col_starts[cid-1]
    col2 = col1 + 1 if cid < 5 else col1

    def merge_or_single(r, c1, c2, val):
        if c1 != c2:
            ws8.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws8.cell(r, c1, val)
        return cell

    cell = merge_or_single(5, col1, col2, f"CLUSTER {cid}")
    cell.fill = fill(hex_clr); cell.font = font("F0F4F8", 9, True)
    cell.alignment = aln("center"); cell.border = border_thin()
    ws8.row_dimensions[5].height = 16

    cell = merge_or_single(6, col1, col2, label)
    cell.fill = fill(NAVY); cell.font = Font(name="Arial", color=hex_clr, size=11, bold=True)
    cell.alignment = aln("center", wrap=True); cell.border = border_thin()
    ws8.row_dimensions[6].height = 20

    cell = merge_or_single(7, col1, col2, f"Dominant: {defn['dominant']}")
    cell.fill = fill(PANEL); cell.font = font(MUTED, 9)
    cell.alignment = aln("center"); cell.border = border_thin()
    ws8.row_dimensions[7].height = 16

    members_str = "\n".join(defn["members"])
    cell = merge_or_single(8, col1, col2, members_str)
    cell.fill = fill("111927"); cell.font = font("F0F4F8", 9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin()
    ws8.row_dimensions[8].height = max(16 * defn["n"], 48)

    cell = merge_or_single(9, col1, col2, desc)
    cell.fill = fill(BLUE); cell.font = font(MUTED, 8)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin()
    ws8.row_dimensions[9].height = 36

    factor_labels = ["ERP","TERM","CREDIT","INFL","LIQ"]
    for fi in range(5):
        r = 10 + fi
        val  = float(defn["centroid"][fi])
        cell = merge_or_single(r, col1, col2, f"{factor_labels[fi]}:  {val:+.2f}")
        bg_c = "1E3A1A" if val > 0.5 else "3A1A1A" if val < -0.5 else PANEL
        cell.fill = fill(bg_c)
        cell.font = Font(name="Arial",
                         color="6BCB77" if val > 0.5 else "FF6B6B" if val < -0.5 else "D0D8E4",
                         size=9, bold=abs(val)>0.5)
        cell.alignment = aln("center"); cell.border = border_thin()
        ws8.row_dimensions[r].height = 14

ws8.row_dimensions[15].height = 12

# Chart 1: Dendrogram
fig1, ax1 = plt.subplots(figsize=(13, 4.5))
dend = dendrogram(link, labels=[dn(a) for a in assets], orientation="top",
                  leaf_rotation=40, leaf_font_size=9,
                  color_threshold=link[-4, 2], ax=ax1)
ax1.set_title("Ward Linkage Dendrogram — Factor Loading Distance\nAssets that merge first are most similar in factor space",
              fontsize=11, fontweight="bold", color="#F0F4F8", pad=10)
ax1.set_ylabel("Ward Distance (lower = more similar)", fontsize=9)
ax1.grid(axis="y", alpha=0.3, linestyle="--")
ax1.tick_params(axis="x", colors="#D0D8E4")
ax1.tick_params(axis="y", colors="#D0D8E4")
plt.tight_layout()
img_dend = img_to_xl(fig1)
img_dend.width = 910; img_dend.height = 340
ws8.add_image(img_dend, "B16")

# Chart 2: Beta heatmap
sorted_idx     = np.argsort(labels_5)
B_sorted       = B_z[sorted_idx]
names_sorted   = [dn(assets[i]) for i in sorted_idx]
cluster_sorted = labels_5[sorted_idx]

fig2, ax2 = plt.subplots(figsize=(7, 6))
im = ax2.imshow(B_sorted, cmap="RdBu_r", vmin=-2.5, vmax=2.5, aspect="auto")
ax2.set_xticks(range(5))
ax2.set_xticklabels(["Equity\nPremium","Term\nPremium","Credit\nSpread","Inflation","Liquidity"],
                     fontsize=9, color="#D0D8E4")
ax2.set_yticks(range(len(assets)))
ax2.set_yticklabels(names_sorted, fontsize=8, color="#D0D8E4")
for i in range(len(assets)):
    for j in range(5):
        ax2.text(j, i, f"{B_sorted[i,j]:.2f}", ha="center", va="center",
                 fontsize=7.5, color="white" if abs(B_sorted[i,j])>1.2 else "#A0B0C0",
                 fontweight="bold" if abs(B_sorted[i,j])>1.5 else "normal")
cbar = plt.colorbar(im, ax=ax2, fraction=0.04, pad=0.02)
cbar.set_label("z-score", color="#D0D8E4", fontsize=9)
plt.setp(cbar.ax.yaxis.get_ticklabels(), color="#D0D8E4")
boundaries = []
prev = cluster_sorted[0]
for i, c in enumerate(cluster_sorted):
    if c != prev:
        boundaries.append(i - 0.5)
        prev = c
for b in boundaries:
    ax2.axhline(y=b, color="white", linewidth=2, alpha=0.9)
ax2.set_title("Standardised Factor Beta Heatmap\nAssets Grouped by Cluster",
              fontsize=10, fontweight="bold", color="#F0F4F8", pad=8)
plt.tight_layout()
img_heat = img_to_xl(fig2)
img_heat.width = 490; img_heat.height = 460
ws8.add_image(img_heat, "B42")

# Chart 3: Cluster centroid profiles
fig3, ax3 = plt.subplots(figsize=(8, 5))
clr_list5 = ["#B03A2E","#2E6DA4","#B7770D","#1E7B45","#6B3FA0"]
x = np.arange(5)
bar_w = 0.15
for cid in range(1, 6):
    centroid = CLUSTER_DEFS[cid]["centroid"]
    offset   = (cid - 3) * bar_w
    ax3.bar(x + offset, centroid, bar_w,
            label=f"C{cid}: {CLUSTER_META[cid][0][:18]}",
            color=clr_list5[cid-1], alpha=0.9)
ax3.axhline(y=0, color="#7A90A8", linewidth=1, linestyle="-")
ax3.set_xticks(x)
ax3.set_xticklabels(["Equity\nPremium","Term\nPremium","Credit\nSpread","Inflation","Liquidity"],
                     fontsize=10, color="#D0D8E4")
ax3.set_ylabel("Mean Standardised Beta (z-score)", fontsize=9, color="#D0D8E4")
ax3.set_title("Cluster Factor Centroid Profiles\nWhat Each Cluster Is Exposed To",
              fontsize=11, fontweight="bold", color="#F0F4F8", pad=10)
ax3.legend(fontsize=8, framealpha=0.3, loc="upper right")
ax3.grid(axis="y", alpha=0.35, linestyle="--")
ax3.tick_params(colors="#D0D8E4")
plt.tight_layout()
img_cent = img_to_xl(fig3)
img_cent.width = 560; img_cent.height = 390
ws8.add_image(img_cent, "G42")

# Full asset detail table
row_tbl = 70
ws8.merge_cells(f"B{row_tbl}:J{row_tbl}")
c = ws8[f"B{row_tbl}"]
c.value = "FULL ASSET DETAIL — CLUSTER ASSIGNMENT AND STANDARDISED BETAS"
c.fill = fill(BLUE); c.font = font("F0F4F8", 11, True)
c.alignment = aln("left"); c.border = border_thin()
ws8.row_dimensions[row_tbl].height = 20

header_row(ws8, row_tbl+1,
    ["Asset Class","Cl.","Cluster Name","ERP (z)","Term (z)","Credit (z)","Inflation (z)","Liquidity (z)","R-Sq"],
    [28, 5, 26, 12, 12, 12, 14, 14, 10])

sort_order = sorted(range(len(assets)), key=lambda i: (labels_5[i], -B_z[i,0]))
for row_i, i in enumerate(sort_order):
    asset = assets[i]
    cid   = int(labels_5[i])
    r2v   = float(result.r_squared.loc[asset]) if asset in result.r_squared.index else 0
    meta  = CLUSTER_META[cid]
    hex_c = meta[1]
    bg    = "1A2436" if row_i % 2 == 0 else PANEL
    row_r = row_tbl + 2 + row_i

    s = Side(style="medium", color=hex_c)
    cell = ws8.cell(row_r, 2, dn(asset))
    cell.fill = fill(bg); cell.font = font("F0F4F8", 10); cell.alignment = aln("left")
    cell.border = Border(left=s, right=Side(style="thin", color="1E2D3D"),
                         top=Side(style="thin", color="1E2D3D"),
                         bottom=Side(style="thin", color="1E2D3D"))

    cell2 = ws8.cell(row_r, 3, cid)
    cell2.fill = fill(PANEL); cell2.font = Font(name="Arial", color=hex_c, size=10, bold=True)
    cell2.alignment = aln("center"); cell2.border = border_thin()

    cell3 = ws8.cell(row_r, 4, meta[0])
    cell3.fill = fill(bg); cell3.font = Font(name="Arial", color=hex_c, size=9)
    cell3.alignment = aln("left"); cell3.border = border_thin()

    for j in range(5):
        val  = float(B_z[i, j])
        cell = ws8.cell(row_r, 5+j, round(val, 3))
        bg_v = "1E3A1A" if val > 0.8 else "3A1A1A" if val < -0.8 else bg
        cell.fill = fill(bg_v)
        cell.font = Font(name="Arial",
                         color="6BCB77" if val > 0.8 else "FF6B6B" if val < -0.8 else "D0D8E4",
                         size=10, bold=abs(val)>1.2)
        cell.alignment = aln("center"); cell.border = border_thin()
        cell.number_format = "+0.000;-0.000;0.000"

    r2_cell = ws8.cell(row_r, 10, r2v)
    r2_cell.fill = fill(bg)
    r2_color = GREEN if r2v >= 0.7 else AMBER if r2v >= 0.4 else RED
    r2_cell.font = Font(name="Arial", color=r2_color, size=10, bold=r2v>=0.7)
    r2_cell.alignment = aln("center"); r2_cell.border = border_thin()
    r2_cell.number_format = "0.0%"
    ws8.row_dimensions[row_r].height = 16

row_note = row_tbl + 2 + len(assets) + 1
ws8.merge_cells(f"B{row_note}:J{row_note}")
c = ws8[f"B{row_note}"]
c.value = ("Green = positive factor tilt  |  Red = negative factor tilt  |  "
           "Bold = strong tilt (|z| > 1.2)  |  R-sq: green >= 0.70, amber >= 0.40, red < 0.40")
c.fill = fill(NAVY); c.font = font(MUTED, 8)
c.alignment = aln("center")
ws8.row_dimensions[row_note].height = 14

# ── Save ───────────────────────────────────────────────────────────
output_path = "analytics_report.xlsx"
wb.save(output_path)
print(f"\nDone. Saved: {output_path}")
print("Sheets: Executive Summary | Backtest Summary | Portfolio Analytics | Factor Model | Look-Through Exposure | Stress Testing | Risk Decomposition | Factor Performance | Covariance | Portfolio Weights | Cluster Analysis")